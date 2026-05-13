#!/usr/bin/env python3
"""Build the NowServingTO corridor-aggregate dataset.

For each of the 12 launch corridors:
  1. Identify matching BIA polygon(s) from business-improvement-areas dataset
  2. Spatial-join parcels.geojson into corridors (point-in-polygon)
  3. Per parcel compute:
       unused_FSI       = max(0, zoneFsi - existing_FSI)
       unused_GFA_m2    = unused_FSI * lotAreaM2
       developer_$      = unused_GFA_m2 * $5000/m² (Toronto multiplex market anchor)
       carbon_tonnes    = embodied(demo) + embodied(rebuild)
       carbon_$         = carbon_tonnes * $185 (social cost of CO2)
  4. Sum per corridor → data/corridors.json
"""
import json, csv, sys, time, io
from pathlib import Path
from datetime import datetime, timezone, timedelta
from shapely.geometry import shape, Point
from shapely.ops import unary_union
from shapely.prepared import prep
from pyproj import Transformer

TODAY = datetime(2026, 5, 12)
WINDOW_12M = TODAY - timedelta(days=365)
WINDOW_90D = TODAY - timedelta(days=90)

# Toronto Dev Apps coords are EPSG:2952 (NAD83(CSRS) / MTM zone 10); reproject to WGS84.
UTM_TO_WGS84 = Transformer.from_crs("EPSG:2952", "EPSG:4326", always_xy=True)

# Mapping: corridor slug → list of City of Toronto 158-neighbourhood names that overlap.
# Used to attribute Persons in Crisis Calls (no lat/lng, only nbhd code).
CORRIDOR_HOODS = {
    'little-jamaica':    ['Oakwood Village (107)', 'Keelesdale-Eglinton West (110)', 'Beechborough-Greenbrook (112)'],
    'west-chinatown':    ['Kensington-Chinatown (78)'],
    'east-chinatown':    ['North Riverdale (68)', 'South Riverdale (70)'],
    'greektown':         ['Danforth (66)', 'Playter Estates-Danforth (67)'],
    'little-italy':      ['Palmerston-Little Italy (80)', 'Trinity-Bellwoods (81)'],
    'little-portugal':   ['Little Portugal (84)'],
    'koreatown':         ['Annex (95)', 'Wychwood (94)'],
    'little-india':      ['Greenwood-Coxwell (65)'],
    'roncesvalles':      ['Roncesvalles (86)'],
    'corso-italia':      ['Corso Italia-Davenport (92)'],
    'kensington-market': ['Kensington-Chinatown (78)'],
    'parkdale':          ['South Parkdale (85)'],
    'thorncliffe-park':  ['Thorncliffe Park (55)'],
    'flemingdon-park':   ['Flemingdon Park (44)'],
    'rexdale':           ['Mount Olive-Silverstone-Jamestown (2)', 'Rexdale-Kipling (4)', 'Elms-Old Rexdale (5)'],
    'markham-lawrence':  ['Bendale-Glen Andrew (133)', 'Bendale South (157)'],
    'eglinton-brimley':  ['Wexford/Maryvale (119)'],
    'jane-finch':        ['Glenfield-Jane Heights (25)', 'Black Creek (24)'],
    'mount-dennis':      ['Mount Dennis (115)'],
    'weston':            ['Weston (113)', 'Weston-Pelham Park (91)'],
    'agincourt':         ['Agincourt North (129)', 'Agincourt South-Malvern West (128)', "Tam O'Shanter-Sullivan (118)"],
    'malvern':           ['Malvern East (132)', 'Malvern West (159)'],
    'pape-village':      ['Danforth East York (59)', 'Playter Estates-Danforth (67)'],
    'regent-park':       ['Regent Park (72)'],
    'moss-park':         ['Moss Park (73)'],
}

# FSA → corridor mapping (for business licence cancellations — FSA-only geocoding).
CORRIDOR_FSAS = {
    'little-jamaica':    {'M6E','M6C','M6B'},
    'west-chinatown':    {'M5T'},
    'east-chinatown':    {'M4M','M4K'},
    'greektown':         {'M4K','M4J'},
    'little-italy':      {'M6G','M6J'},
    'little-portugal':   {'M6J','M6H'},
    'koreatown':         {'M5R','M6G'},
    'little-india':      {'M4L'},
    'roncesvalles':      {'M6R','M6K'},
    'corso-italia':      {'M6E','M6C'},
    'kensington-market': {'M5T'},
    'parkdale':          {'M6K','M6R'},
    # Inner-suburb communities (added v3)
    'thorncliffe-park':  {'M4H'},
    'flemingdon-park':   {'M3C'},
    'rexdale':           {'M9V','M9W'},
    'markham-lawrence':  {'M1H','M1J'},
    'eglinton-brimley':  {'M1J','M1K','M1P'},
    'jane-finch':        {'M3N','M3J'},
    'mount-dennis':      {'M6N','M6M'},
    'weston':            {'M9N','M9P'},
    'agincourt':         {'M1S','M1T','M1V'},
    'malvern':           {'M1B','M1V','M1X'},
    'pape-village':      {'M4J','M4K'},
    'regent-park':       {'M5A'},
    'moss-park':         {'M5A','M5B'},
}
STOREFRONT_CATS = {
    'EATING OR DRINKING ESTABLISHMENT',
    'TAKE-OUT OR RETAIL FOOD ESTABLISHMENT',
    'PERSONAL SERVICES SETTINGS',
    'LAUNDRY PREMISES',
    'SIDEWALK CAFE',
    'PUBLIC HALL',
}

ROOT = Path(__file__).resolve().parent.parent
PARCELS_PATH = ROOT / 'data' / 'parcels.geojson'
BIA_PATH = Path('/tmp/bia.geojson')
OUT_PATH = ROOT / 'data' / 'corridors.json'

# --- Constants ---
CORRIDOR_BUFFER_M = 150              # meters to buffer BIA polygons (catch residential parcels behind commercial frontage)
MARKET_VALUE_PER_M2 = 5000           # CAD per m² of buildable residential GFA (conservative)
DEMO_EMBODIED_CO2_KG_PER_M2 = 700    # kg CO2 from demolishing 1 m² of existing built form
REBUILD_EMBODIED_CO2_KG_PER_M2 = 1500 # kg CO2 from new mid-rise construction per m² of new GFA
SOCIAL_COST_OF_CARBON = 185          # CAD per tonne CO2 (Canadian federal benchmark)
TYPICAL_STOREY_HEIGHT_M = 3.0        # for converting existingMaxBuildingHeightM → storeys

# --- Corridor → BIA name mapping ---
# Multiple BIAs may map to one corridor; we'll union their polygons.
CORRIDORS = [
    # Main-street ethnic commercial corridors (BIA-anchored)
    {"slug": "little-jamaica",       "title": "Eglinton W / Little Jamaica",        "community": "Caribbean / Black",       "bias": ["York-Eglinton", "Eglinton Hill", "Fairbank Village", "Oakwood Village", "Upper Village"], "nbhds": []},
    {"slug": "west-chinatown",       "title": "Spadina-Dundas / West Chinatown",    "community": "Chinese",                  "bias": ["Chinatown"], "nbhds": []},
    {"slug": "east-chinatown",       "title": "Gerrard-Broadview / East Chinatown", "community": "Chinese / Vietnamese",     "bias": ["Broadview Danforth"], "nbhds": []},
    {"slug": "greektown",            "title": "Danforth / Greektown",               "community": "Greek",                    "bias": ["Greektown on the Danforth"], "nbhds": []},
    {"slug": "little-italy",         "title": "College W / Little Italy",           "community": "Italian",                  "bias": ["Little Italy", "College Promenade"], "nbhds": []},
    {"slug": "little-portugal",      "title": "Dundas W / Little Portugal",         "community": "Portuguese",               "bias": ["Little Portugal Toronto", "Dovercourt Village"], "nbhds": []},
    {"slug": "koreatown",            "title": "Bloor / Koreatown",                  "community": "Korean",                   "bias": ["Korea Town"], "nbhds": []},
    {"slug": "little-india",         "title": "Gerrard E / Little India",           "community": "South Asian (Bengali, Tamil, Hindi)", "bias": ["Gerrard India Bazaar"], "nbhds": []},
    {"slug": "roncesvalles",         "title": "Roncesvalles",                       "community": "Polish",                   "bias": ["Roncesvalles Village"], "nbhds": []},
    {"slug": "corso-italia",         "title": "St. Clair W / Corso Italia",         "community": "Italian",                  "bias": ["Corso Italia", "St. Clair Gardens", "Regal Heights Village", "Hillcrest Village"], "nbhds": []},
    {"slug": "kensington-market",    "title": "Kensington Market",                  "community": "Multi-ethnic (historic)",  "bias": ["Kensington Market"], "nbhds": []},
    {"slug": "parkdale",             "title": "Queen W / Parkdale Village",         "community": "Little Tibet / Roma / Filipino / working class", "bias": ["Parkdale Village"], "nbhds": []},
    # Inner-suburb residential communities (Nbhd-anchored, sometimes BIA-augmented)
    {"slug": "thorncliffe-park",     "title": "Thorncliffe Park",                   "community": "Pakistani / Afghan / S Asian", "bias": [], "nbhds": ["Thorncliffe Park"]},
    {"slug": "flemingdon-park",      "title": "Flemingdon Park",                    "community": "S Asian / Filipino / Latin", "bias": [], "nbhds": ["Flemingdon Park"]},
    {"slug": "rexdale",              "title": "Albion / Rexdale",                   "community": "Caribbean / Somali / W African", "bias": ["Albion Islington Square"], "nbhds": ["Mount Olive-Silverstone-Jamestown", "Rexdale-Kipling", "Elms-Old Rexdale"]},
    {"slug": "markham-lawrence",     "title": "Markham / Lawrence",                 "community": "Sri Lankan Tamil",         "bias": ["Cedarbrae Markham Lawrence Village "], "nbhds": ["Bendale-Glen Andrew", "Bendale South"]},
    {"slug": "eglinton-brimley",     "title": "Eglinton / Brimley",                 "community": "Filipino",                 "bias": ["Wexford Heights", "Kennedy Road"], "nbhds": ["Wexford/Maryvale"]},
    {"slug": "jane-finch",           "title": "Jane and Finch",                     "community": "Caribbean / W African / Latin", "bias": [], "nbhds": ["Glenfield-Jane Heights", "Black Creek"]},
    {"slug": "mount-dennis",         "title": "Mount Dennis",                       "community": "Caribbean / E African / multi", "bias": ["Mount Dennis"], "nbhds": ["Mount Dennis"]},
    {"slug": "weston",               "title": "Weston",                             "community": "Caribbean / S Asian / multi", "bias": ["Weston Village"], "nbhds": ["Weston", "Weston-Pelham Park"]},
    {"slug": "agincourt",            "title": "Agincourt / Sheppard E",             "community": "HK Chinese / Tamil",        "bias": ["Sheppard East Village"], "nbhds": ["Agincourt North", "Agincourt South-Malvern West", "Tam O'Shanter-Sullivan"]},
    {"slug": "malvern",              "title": "Malvern",                            "community": "Tamil / Caribbean",        "bias": [], "nbhds": ["Malvern East", "Malvern West"]},
    {"slug": "pape-village",         "title": "Pape Village",                       "community": "Greek / Italian",          "bias": ["Pape Village"], "nbhds": []},
    # Regent Park ONLY (Cabbagetown was previously lumped in here and was crushing the heritage chart by being affluent
    # HCD-protected Victorian rowhouses. The Regent Park post-war housing project is a different, relevant story).
    {"slug": "regent-park",          "title": "Regent Park",                        "community": "Multi-ethnic, historically Black / Bengali / Somali", "bias": [], "nbhds": ["Regent Park"]},
    {"slug": "moss-park",            "title": "Moss Park",                          "community": "Urban Indigenous (largest concentration in TO) · Black Caribbean / Jamaican · multi-ethnic working class · housed census misses ~13× the Indigenous reality", "bias": [], "nbhds": ["Moss Park"]},
]

def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)

# --- Ray-casting point-in-polygon ---
def point_in_ring(pt, ring):
    x, y = pt; inside = False
    n = len(ring)
    for i in range(n):
        x1, y1 = ring[i]; x2, y2 = ring[(i+1) % n]
        if ((y1 > y) != (y2 > y)) and (x < (x2-x1) * (y-y1) / (y2-y1+1e-12) + x1):
            inside = not inside
    return inside

def point_in_polygon(pt, geom):
    """geom: GeoJSON Polygon or MultiPolygon"""
    if geom['type'] == 'Polygon':
        rings = geom['coordinates']
        if not rings: return False
        if not point_in_ring(pt, rings[0]): return False
        for hole in rings[1:]:
            if point_in_ring(pt, hole): return False
        return True
    elif geom['type'] == 'MultiPolygon':
        for poly_rings in geom['coordinates']:
            if not poly_rings: continue
            if point_in_ring(pt, poly_rings[0]):
                inside = True
                for hole in poly_rings[1:]:
                    if point_in_ring(pt, hole): inside = False; break
                if inside: return True
        return False
    return False

def bbox_of(geom):
    pts = []
    if geom['type'] == 'Polygon':
        for ring in geom['coordinates']: pts.extend(ring)
    elif geom['type'] == 'MultiPolygon':
        for poly in geom['coordinates']:
            for ring in poly: pts.extend(ring)
    if not pts: return None
    xs = [p[0] for p in pts]; ys = [p[1] for p in pts]
    return (min(xs), min(ys), max(xs), max(ys))

def in_bbox(pt, bb):
    x, y = pt; x0, y0, x1, y1 = bb
    return x0 <= x <= x1 and y0 <= y <= y1

# --- Step 1: Load BIA polygons + build corridor → list-of-polygons ---
log("Loading BIA polygons…")
with open(BIA_PATH) as f:
    bia_geo = json.load(f)
log(f"  {len(bia_geo['features'])} BIAs loaded")

bia_by_name = {}
for ft in bia_geo['features']:
    name = (ft['properties'].get('AREA_NAME') or '').strip()
    bia_by_name[name] = ft

# Convert buffer from meters to degrees (rough: 1deg ≈ 111km, but at Toronto's latitude 43.7N, lng deg ≈ 80.5km)
BUF_DEG_LAT = CORRIDOR_BUFFER_M / 111_000
BUF_DEG_LNG = CORRIDOR_BUFFER_M / 80_500
# Use the average — shapely buffer is isotropic in degrees; we'll accept the asymmetry
BUF_DEG = (BUF_DEG_LAT + BUF_DEG_LNG) / 2

# Load neighbourhood polygons (city's 158 nbhds) for residential-mode communities
log("Loading 158-neighbourhood polygons…")
with open('/tmp/nbhds.geojson') as f:
    nbhd_geo = json.load(f)
nbhd_by_name = {}
for ft in nbhd_geo['features']:
    n = (ft['properties'].get('AREA_NAME') or '').strip()
    nbhd_by_name[n] = ft
log(f"  {len(nbhd_geo['features'])} nbhd polygons loaded")

for c in CORRIDORS:
    geoms = []
    src_kinds = []
    for bia_name in c.get('bias', []):
        ft = bia_by_name.get(bia_name)
        if not ft:
            log(f"  ⚠ BIA not found: '{bia_name}' for {c['slug']}")
            continue
        geoms.append(shape(ft['geometry'])); src_kinds.append('BIA')
    for nh_name in c.get('nbhds', []):
        ft = nbhd_by_name.get(nh_name)
        if not ft:
            log(f"  ⚠ Nbhd not found: '{nh_name}' for {c['slug']}")
            continue
        geoms.append(shape(ft['geometry'])); src_kinds.append('NBHD')
    if not geoms:
        log(f"  ⚠ NO POLYGONS for corridor '{c['slug']}' — skipping")
        c['_shape'] = None; c['_bbox'] = None; continue
    # For nbhd-only corridors, don't apply the buffer (they're already big polygons covering residential blocks)
    has_bia = 'BIA' in src_kinds
    buf = BUF_DEG if has_bia else 0
    union = unary_union(geoms).buffer(buf)
    c['_shape'] = prep(union)
    c['_bbox'] = union.bounds
    c['_source_kind'] = 'mixed' if (set(src_kinds) == {'BIA','NBHD'}) else ('BIA' if has_bia else 'NBHD')
    log(f"  {c['slug']}: {len(geoms)} {c['_source_kind']} polygon(s), area={union.area*111_000*80_500/1e6:.2f}km²")

# --- Step 2: Stream parcels.geojson, find matches per corridor ---
log(f"Streaming parcels.geojson… ({PARCELS_PATH.stat().st_size//1_000_000} MB)")

# Initialize corridor aggregates
for c in CORRIDORS:
    c.update({
        'parcels': 0,
        'developer_upside_total': 0.0,
        'embodied_carbon_kg_total': 0.0,
        'carbon_cost_total': 0.0,
        'lot_area_m2_total': 0.0,
        'unused_gfa_m2_total': 0.0,
        'existing_gfa_m2_total': 0.0,
        'heritage_designated': 0,
        'parcels_with_zoning': 0,
        '_sample_parcels': [],
    })

# Open and stream
fh = open(PARCELS_PATH)
buf = ''
target = '"features":['
while target not in buf:
    chunk = fh.read(1 << 16)
    if not chunk: break
    buf += chunk
idx = buf.find(target) + len(target)
buf = buf[idx:]

n_total = 0
n_matched = 0
t0 = time.time()

def parse_feature(text):
    """Return a single Feature dict from text starting at first '{'."""
    depth = 0; start = -1
    for i, c in enumerate(text):
        if c == '{':
            if depth == 0: start = i
            depth += 1
        elif c == '}':
            depth -= 1
            if depth == 0 and start != -1:
                return text[start:i+1], i+1
    return None, -1

# Parse features one at a time, growing buf as needed
while True:
    feat_text, consumed = parse_feature(buf)
    if feat_text is None:
        more = fh.read(1 << 16)
        if not more: break
        buf += more
        continue
    buf = buf[consumed:]
    # Skip trailing comma/whitespace
    buf = buf.lstrip(' \n\r\t,')
    if buf.startswith(']'): break

    n_total += 1
    if n_total % 50000 == 0:
        rate = n_total / max(0.1, time.time() - t0)
        log(f"  scanned {n_total:,} parcels ({rate:.0f}/s), matched {n_matched:,}")

    try:
        feat = json.loads(feat_text)
    except Exception:
        continue
    geom = feat.get('geometry')
    if not geom or geom.get('type') != 'Point': continue
    pt = geom['coordinates']
    if not pt or len(pt) < 2: continue
    props = feat.get('properties', {})

    # Quick FSI/lot area sanity
    zone_fsi = props.get('zoneFsi')
    lot_m2 = props.get('lotAreaM2')
    if not (zone_fsi and lot_m2 and zone_fsi > 0 and lot_m2 > 0): continue
    if not props.get('residential', True): pass  # we want ALL parcels in corridors, residential or not

    # Test each corridor's buffered polygon (bbox prefilter, then prepared-geom intersect)
    matched_c = None
    p = Point(pt[0], pt[1])
    for c in CORRIDORS:
        if c['_bbox'] is None: continue
        bx0, by0, bx1, by1 = c['_bbox']
        if not (bx0 <= pt[0] <= bx1 and by0 <= pt[1] <= by1): continue
        if c['_shape'].contains(p):
            matched_c = c
            break
    if not matched_c: continue

    n_matched += 1

    # Per-parcel math
    height_m = props.get('existingMaxBuildingHeightM') or 0
    cov = props.get('buildingCoverageRatio') or 0
    storeys = max(1, height_m / TYPICAL_STOREY_HEIGHT_M) if height_m > 0 else 1
    existing_gfa = cov * lot_m2 * storeys
    existing_fsi = existing_gfa / lot_m2 if lot_m2 > 0 else 0
    unused_fsi = max(0, zone_fsi - existing_fsi)
    unused_gfa = unused_fsi * lot_m2
    developer_upside = unused_gfa * MARKET_VALUE_PER_M2

    # Carbon math (only counts if demo+rebuild happens; show what's at stake)
    demo_co2 = existing_gfa * DEMO_EMBODIED_CO2_KG_PER_M2
    rebuild_co2 = unused_gfa * REBUILD_EMBODIED_CO2_KG_PER_M2
    total_co2_kg = demo_co2 + rebuild_co2
    carbon_cost = (total_co2_kg / 1000) * SOCIAL_COST_OF_CARBON

    matched_c['parcels'] += 1
    matched_c['developer_upside_total'] += developer_upside
    matched_c['embodied_carbon_kg_total'] += total_co2_kg
    matched_c['carbon_cost_total'] += carbon_cost
    matched_c['lot_area_m2_total'] += lot_m2
    matched_c['unused_gfa_m2_total'] += unused_gfa
    matched_c['existing_gfa_m2_total'] += existing_gfa
    if props.get('heritageStatus'):
        matched_c['heritage_designated'] += 1
    matched_c['parcels_with_zoning'] += 1
    if len(matched_c['_sample_parcels']) < 5:
        matched_c['_sample_parcels'].append({
            'parcelId': props.get('parcelId'),
            'address': props.get('address'),
            'zoneFsi': zone_fsi,
            'existingFsiApprox': round(existing_fsi, 2),
            'developerUpsideCAD': round(developer_upside, 0),
            'carbonCostCAD': round(carbon_cost, 0),
        })

fh.close()
log(f"Done streaming: {n_total:,} total parcels, {n_matched:,} matched in 12 corridors")
log(f"Elapsed: {time.time()-t0:.1f}s")

# --- Step 3: Format output ---
out = {
    "generatedAt": datetime.now(timezone.utc).isoformat(),
    "constants": {
        "marketValuePerM2": MARKET_VALUE_PER_M2,
        "demoEmbodiedCO2KgPerM2": DEMO_EMBODIED_CO2_KG_PER_M2,
        "rebuildEmbodiedCO2KgPerM2": REBUILD_EMBODIED_CO2_KG_PER_M2,
        "socialCostOfCarbonCADPerTonne": SOCIAL_COST_OF_CARBON,
        "typicalStoreyHeightM": TYPICAL_STOREY_HEIGHT_M,
    },
    "totals": {
        "parcels": sum(c['parcels'] for c in CORRIDORS),
        "developerUpsideCAD": sum(c['developer_upside_total'] for c in CORRIDORS),
        "carbonCostCAD": sum(c['carbon_cost_total'] for c in CORRIDORS),
        "embodiedCarbonTonnes": sum(c['embodied_carbon_kg_total'] for c in CORRIDORS) / 1000,
    },
    "corridors": []
}
for c in CORRIDORS:
    out['corridors'].append({
        "slug": c['slug'],
        "title": c['title'],
        "community": c.get('community', ''),
        "sourceKind": c.get('_source_kind', 'BIA'),
        "bias": c.get('bias', []),
        "nbhds": c.get('nbhds', []),
        "parcels": c['parcels'],
        "developerUpsideCAD": round(c['developer_upside_total']),
        "carbonCostCAD": round(c['carbon_cost_total']),
        "embodiedCarbonTonnes": round(c['embodied_carbon_kg_total'] / 1000),
        "lotAreaM2": round(c['lot_area_m2_total']),
        "unusedGfaM2": round(c['unused_gfa_m2_total']),
        "existingGfaM2": round(c['existing_gfa_m2_total']),
        "heritageDesignated": c['heritage_designated'],
        "sampleParcels": c['_sample_parcels'],
        # New data layers wired in
        "census": c.get('_census'),
        "shelters": c.get('_shelters'),
        "apt": c.get('_apt'),
        "rentsafe": c.get('_rentsafe'),
    })

# --- Compute RISK INDEX + per-metric rankings ---
# Higher Risk Index = MORE under-pressure / MORE under-protected / MORE exposed.
# Each corridor gets a 0-100 score and a rank 1-12 per metric.
N = len(out['corridors'])

def normalize(vals, invert=False):
    """Min-max normalize to 0-100. invert=True if low-is-bad."""
    vs = [v for v in vals if v is not None]
    if not vs: return [50] * len(vals)
    mn, mx = min(vs), max(vs)
    if mx == mn: return [50] * len(vals)
    out_norm = []
    for v in vals:
        if v is None: out_norm.append(50)
        else:
            n = 100 * (v - mn) / (mx - mn)
            out_norm.append(100 - n if invert else n)
    return out_norm

def rank_high_to_low(vals):
    """Return rank 1..N where 1 = highest value."""
    order = sorted(range(len(vals)), key=lambda i: -vals[i])
    rank = [0] * len(vals)
    for r, idx in enumerate(order, start=1): rank[idx] = r
    return rank

def rank_low_to_high(vals):
    order = sorted(range(len(vals)), key=lambda i: vals[i])
    rank = [0] * len(vals)
    for r, idx in enumerate(order, start=1): rank[idx] = r
    return rank

# Pull metric vectors
upside = [c['developerUpsideCAD'] for c in out['corridors']]
carbon = [c['carbonCostCAD'] for c in out['corridors']]
unused = [c['unusedGfaM2'] for c in out['corridors']]
heritage_per_parcel = [
    (c['heritageDesignated'] / c['parcels']) if c['parcels'] > 0 else 0
    for c in out['corridors']
]
# Risk Index — 50% developer upside, 30% inverse heritage protection, 20% carbon at stake
n_upside = normalize(upside)
n_carbon = normalize(carbon)
n_heritage_inv = normalize(heritage_per_parcel, invert=True)  # low protection = high risk
risk_index = [
    round(0.50 * n_upside[i] + 0.30 * n_heritage_inv[i] + 0.20 * n_carbon[i], 1)
    for i in range(N)
]

# Per-metric ranks
rank_risk     = rank_high_to_low(risk_index)
rank_upside   = rank_high_to_low(upside)
rank_carbon   = rank_high_to_low(carbon)
rank_unused   = rank_high_to_low(unused)
rank_protect  = rank_high_to_low(heritage_per_parcel)  # high heritage/parcel = #1 best protected
rank_unprotected = rank_low_to_high(heritage_per_parcel)  # low heritage/parcel = #1 most under-protected
rank_parcels  = rank_high_to_low([c['parcels'] for c in out['corridors']])

for i, c in enumerate(out['corridors']):
    c['riskIndex'] = risk_index[i]
    c['heritagePerParcel'] = round(heritage_per_parcel[i], 3)
    c['ranks'] = {
        'riskIndex': rank_risk[i],
        'developerUpside': rank_upside[i],
        'carbonCost': rank_carbon[i],
        'unusedGfa': rank_unused[i],
        'mostProtected': rank_protect[i],
        'mostUnprotected': rank_unprotected[i],
        'parcels': rank_parcels[i],
    }

# ---------------------------------------------------------------
# JOIN: Active development applications (last 12 months)
# UTM → WGS84 → corridor polygon
# ---------------------------------------------------------------
log("Joining development applications…")
for c in CORRIDORS: c['_dev_apps'] = 0
n_da = 0; n_da_matched = 0
import re
def parse_iso(s):
    s = (s or '').strip()
    if not s: return None
    try: return datetime.strptime(s[:10], '%Y-%m-%d')
    except: return None

with open('/tmp/dev_apps.csv', encoding='utf-8', errors='replace') as f:
    rdr = csv.DictReader(f)
    for row in rdr:
        n_da += 1
        d = parse_iso(row.get('DATE_SUBMITTED'))
        if not d or d < WINDOW_12M: continue
        try:
            x = float(row['X']); y = float(row['Y'])
        except: continue
        lng, lat = UTM_TO_WGS84.transform(x, y)
        p = Point(lng, lat)
        for c in CORRIDORS:
            if c['_shape'] is None: continue
            bx0, by0, bx1, by1 = c['_bbox']
            if not (bx0 <= lng <= bx1 and by0 <= lat <= by1): continue
            if c['_shape'].contains(p):
                c['_dev_apps'] += 1
                n_da_matched += 1
                break
log(f"  scanned {n_da:,} dev apps, {n_da_matched:,} in corridors (last 12 months)")

# ---------------------------------------------------------------
# JOIN: Major Crime Indicators (last 12 months) via LAT_WGS84/LONG_WGS84
# ---------------------------------------------------------------
log("Joining Major Crime Indicators (last 12 months)…")
for c in CORRIDORS: c['_mci'] = 0
n_mci = 0; n_mci_matched = 0
with open('/tmp/mci.csv', encoding='utf-8', errors='replace') as f:
    rdr = csv.DictReader(f)
    for row in rdr:
        n_mci += 1
        d = parse_iso(row.get('OCC_DATE') or row.get('REPORT_DATE'))
        if not d or d < WINDOW_12M: continue
        try:
            lat = float(row['LAT_WGS84']); lng = float(row['LONG_WGS84'])
        except: continue
        if not lat or not lng: continue
        p = Point(lng, lat)
        for c in CORRIDORS:
            if c['_shape'] is None: continue
            bx0, by0, bx1, by1 = c['_bbox']
            if not (bx0 <= lng <= bx1 and by0 <= lat <= by1): continue
            if c['_shape'].contains(p):
                c['_mci'] += 1
                n_mci_matched += 1
                break
log(f"  scanned {n_mci:,} MCI, {n_mci_matched:,} in corridors (last 12 months)")

# ---------------------------------------------------------------
# JOIN: Persons in Crisis (last 12 months) via 158-nbhd name match
# ---------------------------------------------------------------
log("Joining Persons in Crisis (last 12 months)…")
for c in CORRIDORS: c['_pic'] = 0
# Build reverse map: nbhd_name → corridor slug
hood_to_corridor = {}
for slug, hoods in CORRIDOR_HOODS.items():
    for h in hoods: hood_to_corridor[h] = slug
n_pic = 0; n_pic_matched = 0
with open('/tmp/pic.csv', encoding='utf-8', errors='replace') as f:
    rdr = csv.DictReader(f)
    for row in rdr:
        n_pic += 1
        d = parse_iso(row.get('EVENT_DATE'))
        if not d or d < WINDOW_12M: continue
        hood = (row.get('NEIGHBOURHOOD_158') or '').strip()
        slug = hood_to_corridor.get(hood)
        if not slug: continue
        for c in CORRIDORS:
            if c['slug'] == slug:
                c['_pic'] += 1
                n_pic_matched += 1
                break
log(f"  scanned {n_pic:,} PiC events, {n_pic_matched:,} in corridors (last 12 months)")

# ---------------------------------------------------------------
# JOIN: Business Licence cancellations (last 90 days) via FSA
# ---------------------------------------------------------------
log("Joining business-licence cancellations (last 90 days)…")
for c in CORRIDORS: c['_cancellations_90d'] = 0
def parse_d(s):
    if not s: return None
    s = s.strip()
    for fmt in ("%Y-%m-%d","%d-%b-%Y","%m/%d/%Y"):
        try: return datetime.strptime(s, fmt)
        except: pass
    return None
def fsa_of(s):
    if not s: return None
    m = re.search(r'\bM\d[A-Z]\b', s)
    return m.group(0) if m else None

n_bl = 0; n_bl_matched = 0
with open('/tmp/business_licences_alt.csv', encoding='utf-8', errors='replace') as f:
    rdr = csv.DictReader(f)
    for row in rdr:
        n_bl += 1
        cat = (row.get('Category') or '').strip()
        if cat not in STOREFRONT_CATS: continue
        can = parse_d(row.get('Cancel Date'))
        if not can or can < WINDOW_90D: continue
        fsa = fsa_of(row.get('Licence Address Line 3'))
        if not fsa: continue
        for c in CORRIDORS:
            if fsa in CORRIDOR_FSAS.get(c['slug'], set()):
                c['_cancellations_90d'] += 1
                n_bl_matched += 1
                break
log(f"  scanned {n_bl:,} business licences, {n_bl_matched:,} cancellations matched (last 90d, storefronts)")

# ---------------------------------------------------------------
# CENSUS 2021 — Neighbourhood Profiles (158 nbhd model)
# Pulls tenure, income, indigenous identity, visible minority, immigrant status per nbhd,
# then aggregates across each corridor's constituent nbhds.
# ---------------------------------------------------------------
log("Loading 2021 Census Neighbourhood Profiles…")
import openpyxl as _xl
_wb = _xl.load_workbook('/tmp/nbhd_census.xlsx', read_only=True, data_only=True)
_ws = _wb[_wb.sheetnames[0]]

# Specific row indices in the XLSX (1-based per openpyxl, but iter_rows uses 0-based enumerate)
CENSUS_ROWS = {
    'total_population': 3,
    'median_household_income_after_tax_2020': 246,
    'low_income_lim_at_pct': 178,
    'tenure_total': 299,
    'tenure_owner': 300,
    'tenure_renter': 301,
    'tenure_govt_housing': 302,
    'indigenous_total_pop': 1446,
    'indigenous_identity': 1447,
    'immigrant_total_pop': 1485,
    'immigrants': 1487,
    'recent_immigrants_2011_2021': 1492,
    'non_permanent_residents': 1495,
    'visible_minority_total': 1642,
    'vm_south_asian': 1643,
    'vm_chinese': 1644,
    'vm_black': 1645,
    'vm_filipino': 1646,
    'vm_arab': 1647,
    'vm_latin_american': 1648,
    'vm_southeast_asian': 1649,
    'vm_west_asian': 1650,
    'vm_korean': 1651,
    'vm_japanese': 1652,
    'not_visible_minority': 1655,
}

# Load nbhd column → name map, then load the rows we need
nbhd_col_to_name = {}
census_data = {}  # key → dict[nbhd_name → value]
for i, row in enumerate(_ws.iter_rows(values_only=True)):
    if i == 0:
        # Header: col 0 is "Neighbourhood Name", cols 1..158 are nbhd names
        for ci, v in enumerate(row):
            if ci > 0 and v:
                nbhd_col_to_name[ci] = str(v).strip()
    for key, ridx in CENSUS_ROWS.items():
        if i == ridx:
            d = {}
            for ci, v in enumerate(row):
                if ci > 0 and ci in nbhd_col_to_name and v is not None and v != '':
                    try: d[nbhd_col_to_name[ci]] = float(v)
                    except: pass
            census_data[key] = d
log(f"  loaded {len(census_data)} census variables across {len(nbhd_col_to_name)} nbhds")

# Map corridor → list of CKAN-style nbhd names. CORRIDOR_HOODS uses names with "(NN)" suffix
# but the Census XLSX uses clean names. Build a stripper.
def strip_id(s):
    """'Annex (95)' → 'Annex'; 'Tam O'Shanter-Sullivan (118)' → 'Tam O'Shanter-Sullivan'."""
    import re as _re
    return _re.sub(r'\s*\(\d+\)\s*$', '', s).strip()

# Map corridor → list of clean nbhd names
# Build from BOTH CORRIDOR_HOODS (158-nbhd model labels) and CORRIDORS[*].nbhds (clean names already)
corridor_to_nbhds = {}
for c in CORRIDORS:
    slug = c['slug']
    clean = []
    # From CORRIDOR_HOODS (PIC mapping)
    for raw in CORRIDOR_HOODS.get(slug, []):
        clean.append(strip_id(raw))
    # From CORRIDORS[*].nbhds (already clean)
    for raw in c.get('nbhds', []):
        if raw not in clean: clean.append(raw)
    corridor_to_nbhds[slug] = clean

# Aggregate census per corridor
def sum_for_corridor(key, nbhd_names):
    total = 0; n = 0
    d = census_data.get(key, {})
    for nb in nbhd_names:
        if nb in d:
            total += d[nb]; n += 1
    return total if n > 0 else None

def weighted_avg_for_corridor(rate_key, denom_key, nbhd_names):
    """For % values, compute population-weighted average."""
    num_total = 0; den_total = 0
    rd = census_data.get(rate_key, {})
    dd = census_data.get(denom_key, {})
    for nb in nbhd_names:
        if nb in rd and nb in dd and dd[nb] > 0:
            num_total += rd[nb] * dd[nb] / 100  # rate × denom = absolute
            den_total += dd[nb]
    return (num_total / den_total * 100) if den_total > 0 else None

log("Aggregating census per corridor…")
for c in CORRIDORS:
    nbs = corridor_to_nbhds.get(c['slug'], [])
    if not nbs:
        c['_census'] = None; continue
    total_pop = sum_for_corridor('total_population', nbs)
    income_med = sum_for_corridor('median_household_income_after_tax_2020', nbs)
    # median values can't be summed; use simple mean across nbhds (best we can do without record-level data)
    income_med_vals = [census_data['median_household_income_after_tax_2020'].get(nb) for nb in nbs
                       if nb in census_data.get('median_household_income_after_tax_2020', {})]
    income_med = sum(income_med_vals) / len(income_med_vals) if income_med_vals else None
    tenure_total = sum_for_corridor('tenure_total', nbs)
    tenure_owner = sum_for_corridor('tenure_owner', nbs)
    tenure_renter = sum_for_corridor('tenure_renter', nbs)
    tenure_govt = sum_for_corridor('tenure_govt_housing', nbs)
    indigenous = sum_for_corridor('indigenous_identity', nbs)
    immigrant_total_pop = sum_for_corridor('immigrant_total_pop', nbs)
    immigrants = sum_for_corridor('immigrants', nbs)
    recent_imm = sum_for_corridor('recent_immigrants_2011_2021', nbs)
    vm_total_pop = sum_for_corridor('visible_minority_total', nbs)
    def pct(num, den): return round(num*100/den, 1) if (num and den and den > 0) else None
    c['_census'] = {
        'totalPopulation': int(total_pop) if total_pop else None,
        'medianHouseholdIncomeAfterTax2020': int(income_med) if income_med else None,
        'lowIncomeLimAtPct': weighted_avg_for_corridor('low_income_lim_at_pct', 'total_population', nbs),
        'ownerPct': pct(tenure_owner, tenure_total),
        'renterPct': pct(tenure_renter, tenure_total),
        'govtHousingPct': pct(tenure_govt, tenure_total),
        'indigenousPct': pct(indigenous, sum_for_corridor('indigenous_total_pop', nbs)),
        'immigrantPct': pct(immigrants, immigrant_total_pop),
        'recentImmigrantPct': pct(recent_imm, immigrant_total_pop),
        'nbhdCount': len([nb for nb in nbs if nb in census_data.get('total_population', {})]),
        'visibleMinorityBreakdown': {
            'south_asian': pct(sum_for_corridor('vm_south_asian', nbs), vm_total_pop),
            'chinese': pct(sum_for_corridor('vm_chinese', nbs), vm_total_pop),
            'black': pct(sum_for_corridor('vm_black', nbs), vm_total_pop),
            'filipino': pct(sum_for_corridor('vm_filipino', nbs), vm_total_pop),
            'arab': pct(sum_for_corridor('vm_arab', nbs), vm_total_pop),
            'latin_american': pct(sum_for_corridor('vm_latin_american', nbs), vm_total_pop),
            'southeast_asian': pct(sum_for_corridor('vm_southeast_asian', nbs), vm_total_pop),
            'korean': pct(sum_for_corridor('vm_korean', nbs), vm_total_pop),
            'west_asian': pct(sum_for_corridor('vm_west_asian', nbs), vm_total_pop),
            'japanese': pct(sum_for_corridor('vm_japanese', nbs), vm_total_pop),
        }
    }
log(f"  census aggregated for {sum(1 for c in CORRIDORS if c.get('_census'))} corridors")

# ---------------------------------------------------------------
# SHELTER OCCUPANCY — daily, group by postal-code FSA, get most-recent snapshot
# ---------------------------------------------------------------
log("Loading Daily Shelter Occupancy…")
shelter_latest = {}  # (shelter_id, location_id) → most-recent record
with open('/tmp/shelter.csv', encoding='utf-8', errors='replace') as f:
    rdr = csv.DictReader(f)
    for row in rdr:
        key = (row.get('SHELTER_ID',''), row.get('LOCATION_ID',''))
        d = row.get('OCCUPANCY_DATE','')
        prev = shelter_latest.get(key)
        if prev is None or d > prev.get('OCCUPANCY_DATE',''):
            shelter_latest[key] = row
log(f"  {len(shelter_latest)} unique shelter-location combos; latest snapshot")
# Aggregate by FSA
shelter_by_fsa = {}
for row in shelter_latest.values():
    pc = (row.get('LOCATION_POSTAL_CODE') or '').strip()
    if not pc: continue
    fsa = pc[:3].upper()
    sec = (row.get('SECTOR') or '').strip()
    try: occupied = int(row.get('OCCUPIED_BEDS') or row.get('OCCUPIED_ROOMS') or 0)
    except: occupied = 0
    if fsa not in shelter_by_fsa:
        shelter_by_fsa[fsa] = {'beds': 0, 'shelters': 0, 'indigenous_specific': 0, 'sectors': set()}
    shelter_by_fsa[fsa]['beds'] += occupied
    shelter_by_fsa[fsa]['shelters'] += 1
    shelter_by_fsa[fsa]['sectors'].add(sec)
    if 'indigenous' in sec.lower(): shelter_by_fsa[fsa]['indigenous_specific'] += 1
log(f"  shelter data aggregated to {len(shelter_by_fsa)} FSAs")

# ---------------------------------------------------------------
# APARTMENT BUILDING REGISTRATION + RENTSAFETO — join by WARD (apt/rentsafe have ward not FSA/postal)
# ---------------------------------------------------------------
# Ward → corridor mapping. Toronto's 25 wards; each corridor sits in 1-2.
CORRIDOR_WARDS = {
    'little-jamaica':   {8, 9},     # Eglinton-Lawrence + Davenport
    'west-chinatown':   {10},        # Spadina-Fort York
    'east-chinatown':   {14},        # Toronto-Danforth
    'greektown':        {14},        # Toronto-Danforth
    'little-italy':     {9, 10},    # Davenport + Spadina-Fort York
    'little-portugal':  {9},         # Davenport
    'koreatown':        {11},        # University-Rosedale
    'little-india':     {14},        # Toronto-Danforth
    'roncesvalles':     {4},         # Parkdale-High Park
    'corso-italia':     {9},         # Davenport
    'kensington-market':{10},        # Spadina-Fort York
    'parkdale':         {4},         # Parkdale-High Park
    'thorncliffe-park': {15},        # Don Valley West
    'flemingdon-park':  {15, 16},    # Don Valley West + Don Valley East
    'rexdale':          {1, 2},      # Etobicoke North
    'markham-lawrence': {21, 24},    # Scarborough Centre / Scarborough-Guildwood
    'eglinton-brimley': {21, 22},    # Scarborough Centre / Scarborough-Agincourt
    'jane-finch':       {6, 7},      # York Centre / Humber River-Black Creek
    'mount-dennis':     {5, 6},      # York South-Weston / York Centre
    'weston':           {5, 7},      # York South-Weston / Humber River-Black Creek
    'agincourt':        {22, 23},    # Scarborough-Agincourt / Scarborough North
    'malvern':          {23, 25},    # Scarborough North / Scarborough-Rouge Park
    'pape-village':     {14},        # Toronto-Danforth
    'regent-park':      {13},        # Toronto Centre
    'moss-park':        {13},        # Toronto Centre
}

log("Loading Apartment Building Registration (by ward)…")
apt_by_ward = {}
with open('/tmp/aptreg.csv', encoding='utf-8', errors='replace') as f:
    rdr = csv.DictReader(f)
    n_apt = 0
    for row in rdr:
        n_apt += 1
        try: ward = int(row.get('WARD') or 0)
        except: ward = 0
        if not ward: continue
        try: units = int(row.get('CONFIRMED_UNITS') or 0)
        except: units = 0
        try: storeys = int(row.get('CONFIRMED_STOREYS') or 0)
        except: storeys = 0
        if ward not in apt_by_ward:
            apt_by_ward[ward] = {'buildings': 0, 'units': 0, '_st_tot': 0}
        apt_by_ward[ward]['buildings'] += 1
        apt_by_ward[ward]['units'] += units
        apt_by_ward[ward]['_st_tot'] += storeys
    for w in apt_by_ward:
        b = apt_by_ward[w]
        b['avg_storeys'] = round(b['_st_tot']/b['buildings'], 1) if b['buildings'] else 0
        del b['_st_tot']
log(f"  scanned {n_apt:,} apt buildings, aggregated to {len(apt_by_ward)} wards")

log("Loading RentSafeTO evaluations (by ward)…")
rs_by_ward = {}
with open('/tmp/rentsafe.csv', encoding='utf-8', errors='replace') as f:
    rdr = csv.DictReader(f)
    n_rs = 0
    for row in rdr:
        n_rs += 1
        try: ward = int(row.get('WARD') or 0)
        except: ward = 0
        if not ward: continue
        try: score = float(row.get('CURRENT BUILDING EVAL SCORE') or 0)
        except: score = 0
        if ward not in rs_by_ward:
            rs_by_ward[ward] = {'buildings': 0, '_score_tot': 0}
        rs_by_ward[ward]['buildings'] += 1
        rs_by_ward[ward]['_score_tot'] += score
    for w in rs_by_ward:
        b = rs_by_ward[w]
        b['avg_score'] = round(b['_score_tot']/b['buildings'], 1) if b['buildings'] else 0
        del b['_score_tot']
log(f"  scanned {n_rs:,} evals, aggregated to {len(rs_by_ward)} wards")

# ---------------------------------------------------------------
# BUSINESS LICENCES — per-corridor breakdown by CATEGORY (active + legacy)
# ---------------------------------------------------------------
log("Building per-corridor business-category breakdown…")
BIZ_CATS = {
    'EATING OR DRINKING ESTABLISHMENT': 'restaurant',
    'TAKE-OUT OR RETAIL FOOD ESTABLISHMENT': 'food_retail',
    'PERSONAL SERVICES SETTINGS': 'personal_services',
    'LAUNDRY PREMISES': 'laundry',
    'SIDEWALK CAFE': 'sidewalk_cafe',
    'PUBLIC HALL': 'public_hall',
    'EATING ESTABLISHMENT': 'restaurant',  # historic name
    'RETAIL STORE (FOOD)': 'food_retail',  # historic name
}
biz_breakdown = {slug: {'total_active': 0, 'legacy_20y': 0, 'legacy_10y': 0, 'by_category': {}, 'by_category_legacy20': {}} for slug in CORRIDOR_FSAS}
with open('/tmp/business_licences_alt.csv', encoding='utf-8', errors='replace') as f:
    rdr = csv.DictReader(f)
    for row in rdr:
        cat = (row.get('Category') or '').strip()
        slim_cat = BIZ_CATS.get(cat)
        if not slim_cat: continue
        if (row.get('Cancel Date') or '').strip(): continue  # active only
        iss = parse_d(row.get('Issued'))
        if not iss: continue
        years = (TODAY - iss).days / 365.25
        fsa = fsa_of(row.get('Licence Address Line 3'))
        if not fsa: continue
        for slug, fsas in CORRIDOR_FSAS.items():
            if fsa in fsas:
                b = biz_breakdown[slug]
                b['total_active'] += 1
                b['by_category'][slim_cat] = b['by_category'].get(slim_cat, 0) + 1
                if years >= 20:
                    b['legacy_20y'] += 1
                    b['by_category_legacy20'][slim_cat] = b['by_category_legacy20'].get(slim_cat, 0) + 1
                if years >= 10:
                    b['legacy_10y'] += 1
                break
log(f"  business breakdown built for {sum(1 for v in biz_breakdown.values() if v['total_active'])} corridors")

# ---------------------------------------------------------------
# CUISINE / ETHNIC ATTRIBUTION via OPERATING NAME pattern matching
# Inferred (cuisine ≠ owner ethnicity), but a real signal where it lands.
# ---------------------------------------------------------------
CUISINE_PATTERNS = {
    'caribbean':     ['ROTI','JERK','CARIBBEAN','JAMAICAN','JAMAICA','ACKEE','PATTY','OXTAIL','PLANTAIN','RASTAFAR','RIDDIM','IRIE','RUDIE','REGGAE','ISLAND'],
    'south_asian':   ['TANDOORI','MASALA','BIRYANI','TIKKA','NAAN','BHATURA','PANEER','DOSA','IDLI','PUNJABI','KARAHI','SAMOSA','KABAB','KEBAB','INDIA','INDIAN','BHOJAN','THALI'],
    'chinese':       ['WOK','CHINESE','CHINA','DIM SUM','SZECHUAN','SICHUAN','HUNAN','CANTONESE','HONG KONG','CHOPSTICK','BAO','BUBBLE TEA','MANDARIN','HOUSE OF NOODLE','MISTER WOK','BAMBOO','MAJESTIC'],
    'vietnamese':    ['PHO','BANH MI','BUN ','VIETNAMESE','SAIGON','HANOI','VIETNAM'],
    'korean':        ['KIMCHI','KOREAN','BIBIMBAP','GANGNAM','SEOUL','HANSAM'],
    'italian':       ['PIZZA','PIZZERIA','PASTA','RISTORANTE','TRATTORIA','GELATERIA','GELATO','ITALIANO','ITALIA','NAPOLI','MILANO','VERONA','TOSCANA'],
    'portuguese':    ['PADARIA','PORTUGUESA','PORTUGUESE','PASTEL','BACALAU','PORTUGAL','LISBOA','PORTO'],
    'greek':         ['SOUVLAKI','GYRO','GREEK','HELLENIC','ATHENS','OLYMPIA','MYKONOS','SANTORINI','ZORBA','KEFI'],
    'japanese':      ['SUSHI','RAMEN','IZAKAYA','JAPANESE','TOKYO','OSAKA','SAKURA','TERIYAKI','SAKE','TONKATSU','UDON','SOBA'],
    'filipino':      ['ADOBO','LECHON','KAINAN','FILIPINO','PINOY','MANILA','PINAS','TAGALOG','SARISAR','TANGKE'],
    'tibetan':       ['MOMO','TIBETAN','TIBET','LHASA','HIMALAY','SHANGRI'],
    'african_horn':  ['ETHIOPIAN','INJERA','ERITREAN','SOMALI','HABESHA','ADDIS','ASMARA','MOGADISHU','HARGEISA'],
    'african_west':  ['NIGERIAN','GHANAIAN','SENEGAL','MALI','AFRICAN','SUYA','JOLLOF','EGUSI','FUFU'],
    'latin':         ['TACO','TAQUERIA','EMPANADA','SALVADOR','LATINO','LATINA','MEXICAN','MEXICO','PERUVIAN','COLOMBIAN','VENEZUELAN','PUPUSAS','CHURRO','ASADO'],
    'polish':        ['POLSKI','POLSKA','PIEROGI','POLISH','KRAKOW','WARSZAW','KIELBASA'],
    'middle_east':   ['SHAWARMA','FALAFEL','LEBANESE','SYRIAN','ARABIAN','PERSIAN','IRANIAN','TURKISH','ANATOLIA','BEIRUT','MEDITERRAN'],
    'tamil':         ['TAMIL','EELAM','JAFFNA','CHENNAI','SRI ','MADRAS','KOTHU'],
    'irish_uk':      ['IRISH','DUBLIN','CELTIC','KILKENNY','SCOTTISH','HIGHLAND','LONDON','BRITISH','CHIPS & FISH','FISH & CHIPS','PUB'],
    'french':        ['FRENCH','BISTRO','BRASSERIE','BOULANGERIE','PATISSERIE','CHEZ ','LE PARIS','LYON','MARSEILLE','CROISSANT'],
    'german':        ['GERMAN','BIERGARTEN','WURST','BAVARIA','SCHWARZWALD','OKTOBER','BRATWURST','SCHNITZEL'],
    'jewish_deli':   ['KOSHER','BAGEL','KNISH','SHTETL','SHWARTZ','UNITED BAKERS','MATZO','YIDDISH','SCHWARTZS'],
    'eastern_eu':    ['UKRAINIAN','RUSSIAN','BULGARIAN','HUNGARIAN','ROMANIAN','BORSCHT','PEROGY','PYROGY','VARENY','KYIV','KIEV','ODESA','PRAGUE','GOULASH','CZECH'],
}
CUISINE_LABEL = {
    'italian':'Italian','chinese':'Chinese','japanese':'Japanese','korean':'Korean',
    'vietnamese':'Vietnamese','filipino':'Filipino','thai':'Thai',
    'indonesian':'Indonesian','malaysian':'Malaysian','burmese':'Burmese',
    'south_asian':'South Asian','indian':'Indian','pakistani':'Pakistani','afghan':'Afghan',
    'bangladeshi':'Bangladeshi','tamil':'Tamil','tibetan':'Tibetan',
    'caribbean':'Caribbean','jamaican':'Jamaican','trinidadian':'Trinidadian','guyanese':'Guyanese','haitian':'Haitian',
    'greek':'Greek','portuguese':'Portuguese','polish':'Polish','french':'French',
    'irish_uk':'Irish/UK','german':'German','jewish_deli':'Jewish deli',
    'eastern_eu':'Eastern European','ukrainian':'Ukrainian','russian':'Russian','hungarian':'Hungarian',
    'middle_east':'Middle Eastern','lebanese':'Lebanese','turkish':'Turkish','syrian':'Syrian','persian':'Persian',
    'latin':'Latin American','mexican':'Mexican','salvadoran':'Salvadoran','peruvian':'Peruvian','colombian':'Colombian','brazilian':'Brazilian',
    'african_horn':'East African','ethiopian':'Ethiopian','eritrean':'Eritrean','somali':'Somali',
    'african_west':'West African','nigerian':'Nigerian','ghanaian':'Ghanaian','moroccan':'Moroccan',
}
log("Tagging businesses by cuisine pattern (inferred from Operating Name)…")
cuisine_by_corridor = {slug: {} for slug in CORRIDOR_FSAS}
n_tagged = 0; n_total_food = 0
with open('/tmp/business_licences_alt.csv', encoding='utf-8', errors='replace') as f:
    rdr = csv.DictReader(f)
    for row in rdr:
        cat = (row.get('Category') or '').strip()
        if cat not in ('EATING OR DRINKING ESTABLISHMENT','TAKE-OUT OR RETAIL FOOD ESTABLISHMENT','EATING ESTABLISHMENT','RETAIL STORE (FOOD)'):
            continue
        if (row.get('Cancel Date') or '').strip(): continue  # active only
        n_total_food += 1
        op = (row.get('Operating Name') or '').upper()
        if not op: continue
        # Match cuisines
        matched = None
        for cuisine, keys in CUISINE_PATTERNS.items():
            for k in keys:
                if k in op: matched = cuisine; break
            if matched: break
        if not matched: continue
        n_tagged += 1
        fsa = fsa_of(row.get('Licence Address Line 3'))
        if not fsa: continue
        for slug, fsas in CORRIDOR_FSAS.items():
            if fsa in fsas:
                cuisine_by_corridor[slug][matched] = cuisine_by_corridor[slug].get(matched, 0) + 1
                break
log(f"  tagged {n_tagged:,} of {n_total_food:,} food businesses ({n_tagged*100/max(n_total_food,1):.0f}%) by cuisine")

# Attach cuisine breakdown to each CORRIDOR for output
for c in CORRIDORS:
    c['_cuisine'] = cuisine_by_corridor.get(c['slug'], {})

# Attach shelter/apt/rentsafe data to corridors via their FSA set
log("Attaching shelter/apt/rentsafe/biz to corridors…")
for c in CORRIDORS:
    slug = c['slug']
    fsas = CORRIDOR_FSAS.get(slug, set())
    wards = CORRIDOR_WARDS.get(slug, set())
    # Shelters (FSA-based, since postal-coded)
    beds = sum(shelter_by_fsa.get(f, {}).get('beds', 0) for f in fsas)
    shelters_count = sum(shelter_by_fsa.get(f, {}).get('shelters', 0) for f in fsas)
    indig_shelters = sum(shelter_by_fsa.get(f, {}).get('indigenous_specific', 0) for f in fsas)
    c['_shelters'] = {'beds_occupied': beds, 'shelters': shelters_count, 'indigenous_specific_shelters': indig_shelters}
    # Apt registrations (ward-based, since SITE_ADDRESS lacks postal)
    apt_bldg = sum(apt_by_ward.get(w, {}).get('buildings', 0) for w in wards)
    apt_units = sum(apt_by_ward.get(w, {}).get('units', 0) for w in wards)
    storeys_vals = [apt_by_ward.get(w, {}).get('avg_storeys', 0) for w in wards if apt_by_ward.get(w, {}).get('buildings', 0) > 0]
    avg_storeys = round(sum(storeys_vals) / len(storeys_vals), 1) if storeys_vals else None
    c['_apt'] = {'registered_buildings': apt_bldg, 'registered_units': apt_units, 'avg_storeys': avg_storeys, 'note': 'ward-aggregated'}
    # RentSafeTO (ward-based)
    rs_bldg = sum(rs_by_ward.get(w, {}).get('buildings', 0) for w in wards)
    rs_scores = [rs_by_ward.get(w, {}).get('avg_score', 0) for w in wards if rs_by_ward.get(w, {}).get('buildings', 0) > 0]
    rs_avg = round(sum(rs_scores) / len(rs_scores), 1) if rs_scores else None
    c['_rentsafe'] = {'evaluated_buildings': rs_bldg, 'avg_score': rs_avg, 'note': 'ward-aggregated'}
    # Business breakdown
    c['_biz'] = biz_breakdown.get(slug, {})

# ---------------------------------------------------------------
# HERITAGE REGISTER — proper spatial join (the parcels.geojson heritageStatus field is stale)
# ---------------------------------------------------------------
log("Spatial-joining Heritage Register designations (proper count, supersedes parcels.geojson value)…")
import zipfile as _zf
import shapefile as _sf
for c in CORRIDORS:
    c['_heritage_proper'] = 0
    c['_heritage_addresses'] = []
with _zf.ZipFile('/tmp/heritage_register.zip') as zf:
    shp = [m for m in zf.infolist() if m.filename.lower().endswith('.shp')][0]
    dbf = [m for m in zf.infolist() if m.filename.lower().endswith('.dbf')][0]
    sfr = _sf.Reader(shp=io.BytesIO(zf.read(shp)), dbf=io.BytesIO(zf.read(dbf)))
    flds = [f[0] for f in sfr.fields if f[0] != 'DeletionFlag']
    n_pts = 0; n_matched = 0
    for sr in sfr.shapeRecords():
        if not sr.shape.points: continue
        n_pts += 1
        lng, lat = sr.shape.points[0]
        p = Point(lng, lat)
        rec = dict(zip(flds, sr.record))
        for c in CORRIDORS:
            if c['_shape'] is None: continue
            bx0, by0, bx1, by1 = c['_bbox']
            if not (bx0 <= lng <= bx1 and by0 <= lat <= by1): continue
            if c['_shape'].contains(p):
                c['_heritage_proper'] += 1
                addr = (rec.get('ADDRESS') or '').strip()
                status = (rec.get('STATUS') or '').strip()
                desc = (rec.get('DESCRIPTIO') or '').strip()
                if len(c['_heritage_addresses']) < 12:
                    c['_heritage_addresses'].append({'address': addr, 'status': status, 'description': desc[:160]})
                n_matched += 1
                break
log(f"  {n_matched} of {n_pts} heritage points matched into 24 corridors")

# ---------------------------------------------------------------
# RECENT CANCELLATIONS FEED — named businesses, last 90 days, storefronts only
# ---------------------------------------------------------------
log("Building cancellations feed…")
cancel_feed = []
with open('/tmp/business_licences_alt.csv', encoding='utf-8', errors='replace') as f:
    rdr = csv.DictReader(f)
    for row in rdr:
        cat = (row.get('Category') or '').strip()
        if cat not in STOREFRONT_CATS: continue
        can = parse_d(row.get('Cancel Date'))
        if not can or can < WINDOW_90D: continue
        iss = parse_d(row.get('Issued'))
        fsa = fsa_of(row.get('Licence Address Line 3'))
        if not fsa: continue
        # Find corridor by FSA (first match wins)
        corridor = None
        for slug, fsas in CORRIDOR_FSAS.items():
            if fsa in fsas: corridor = slug; break
        if not corridor: continue
        op = (row.get('Operating Name') or '').strip()
        addr = (row.get('Licence Address Line 1') or '').strip()
        years = round((can - iss).days / 365.25, 1) if iss else None
        cancel_feed.append({
            'operatingName': op or '(unnamed business)',
            'address': addr,
            'corridor': corridor,
            'category': cat,
            'cancelDate': can.strftime('%Y-%m-%d'),
            'yearsOperating': years,
        })
# Sort: longest-operating first (more newsworthy)
cancel_feed.sort(key=lambda r: -(r['yearsOperating'] or 0))
out['recentCancellations'] = cancel_feed[:30]
log(f"  {len(cancel_feed)} named cancellations in last 90 days; keeping top-30 by years operating")

# ---------------------------------------------------------------
# CUISINE CLOSURE INDEX — citywide cuisine-tagged closures (not BIA-bounded).
# Same classifier as the active-business tagger (CUISINE_PATTERNS above), applied
# to the cancelled-licence feed. Produces a ranked leaderboard + named-ledger
# for the "what's actually closing" homepage section.
# ---------------------------------------------------------------
log("Building cuisine closure index (citywide)…")
FOOD_CATS_FOR_CUISINE = ('EATING OR DRINKING ESTABLISHMENT','TAKE-OUT OR RETAIL FOOD ESTABLISHMENT','EATING ESTABLISHMENT','RETAIL STORE (FOOD)')
WINDOW_1Y_FOOD = TODAY - timedelta(days=365)
cc_closed_90d = {}; cc_closed_1y = {}; cc_active = {}
cc_named_90d = []
n_food_total = 0; n_food_cancelled_90d_all = 0; n_tagged_90d_all = 0
with open('/tmp/business_licences_alt.csv', encoding='utf-8', errors='replace') as f:
    rdr = csv.DictReader(f)
    for row in rdr:
        cat = (row.get('Category') or '').strip()
        if cat not in FOOD_CATS_FOR_CUISINE: continue
        n_food_total += 1
        op_raw = (row.get('Operating Name') or '').strip()
        op = op_raw.upper()
        if not op: continue
        matched = None
        for cuisine_k, keys in CUISINE_PATTERNS.items():
            for kw in keys:
                if kw in op: matched = cuisine_k; break
            if matched: break
        can = parse_d(row.get('Cancel Date'))
        if can:
            if can >= WINDOW_1Y_FOOD and matched:
                cc_closed_1y[matched] = cc_closed_1y.get(matched, 0) + 1
            if can >= WINDOW_90D:
                n_food_cancelled_90d_all += 1
                if matched:
                    cc_closed_90d[matched] = cc_closed_90d.get(matched, 0) + 1
                    n_tagged_90d_all += 1
                    iss = parse_d(row.get('Issued'))
                    yrs = round((can - iss).days / 365.25) if iss else None
                    addr1 = (row.get('Licence Address Line 1') or '').strip()
                    addr3 = (row.get('Licence Address Line 3') or '').strip()
                    cc_named_90d.append({
                        'operatingName': op_raw,
                        'cuisine': matched,
                        'yearsOperating': yrs,
                        'cancelDate': can.strftime('%Y-%m-%d'),
                        'address': (addr1 + ' ' + addr3).strip(),
                    })
        elif matched:
            cc_active[matched] = cc_active.get(matched, 0) + 1
cuisines_out = []
for c in set(list(cc_closed_90d) + list(cc_closed_1y) + list(cc_active)):
    n90 = cc_closed_90d.get(c, 0); n1y = cc_closed_1y.get(c, 0); act = cc_active.get(c, 0)
    cuisines_out.append({
        'key': c, 'closed90d': n90, 'closed1y': n1y, 'active': act,
        'rate1y': round(n1y / act * 100, 1) if act >= 20 else None,
    })
cuisines_out.sort(key=lambda r: -r['closed90d'])
cc_named_90d.sort(key=lambda r: -(r['yearsOperating'] or 0))
# Dedupe identical name+date (chain closures sometimes hit on the same day)
_seen = set(); cc_named_dedup = []
for n in cc_named_90d:
    k = (n['operatingName'], n['cancelDate'])
    if k in _seen: continue
    _seen.add(k); cc_named_dedup.append(n)
out['closuresByCuisine'] = {
    'asOf': TODAY.strftime('%Y-%m-%d'),
    'windowDays': 90,
    'totalFoodScanned': n_food_total,
    'totalFoodCancelled90d': n_food_cancelled_90d_all,
    'tagRate90d': round(n_tagged_90d_all / max(n_food_cancelled_90d_all, 1) * 100, 1),
    'cuisines': cuisines_out,
    'named': cc_named_dedup[:50],
}
log(f"  cuisine closure index: {len(cuisines_out)} cuisines, {len(cc_named_dedup)} named (tag rate {out['closuresByCuisine']['tagRate90d']}%)")

# ---------------------------------------------------------------
# NEW OPENINGS · "Now open" — citywide cuisine-tagged restaurants licensed
# in the last 365 days (Issued recent, no Cancel Date). Positive-frame
# counterpart to the closure index; this is what makes the page returnable.
# ---------------------------------------------------------------
log("Building new-openings index…")
WINDOW_365 = TODAY - timedelta(days=365)
WINDOW_30  = TODAY - timedelta(days=30)

# Load LLM cuisine cache (populated by tools/llm_classify.py + llm_classify_batch.py).
# Trusts LLM>keyword: LLM=unknown drops the entry; LLM=cuisine_X overrides any keyword tag.
LLM_CACHE_PATH = ROOT / 'tools' / 'cache' / 'llm_cuisine_cache.json'
_llm_cache = {}
if LLM_CACHE_PATH.exists():
    _llm_cache = json.loads(LLM_CACHE_PATH.read_text())
    log(f"  loaded LLM cuisine cache: {len(_llm_cache)} entries")
_VALID_LLM_CUISINES = set(CUISINE_LABEL.keys())  # every key with a display label is valid
CUISINE_LABEL.setdefault('thai', 'Thai')

# Load web_verify_cache so cuisine determined from web_search trumps name-only LLM.
_WEB_VERIFY_PATH = ROOT / 'tools' / 'cache' / 'web_verify_cache.json'
_web_verify_cache = {}
if _WEB_VERIFY_PATH.exists():
    _web_verify_cache = json.loads(_WEB_VERIFY_PATH.read_text())

# Chain denylist — substring match against UPPERCASE operating name. Forces None.
_CHAIN_DENYLIST = (
    'POPEYES','POPEYE\'S','KFC','CHURCH\'S CHICKEN','CHURCHS CHICKEN','MARY BROWN',
    'WENDY','BURGER KING','MCDONALD','HARVEY','A&W','TIM HORTON','COFFEE TIME',
    'SECOND CUP','STARBUCKS','TIMOTHY\'S COFFEE',
    'SUBWAY','MR. SUB','MR SUB','QUIZNOS','EXTREME PITA','PITA PIT',
    'APPLEBEE','OUTBACK','IHOP','DENNY','JACK ASTOR','SCORES','KELSEY',
    'MONTANA','EAST SIDE MARIO','BOSTON PIZZA','PIZZA NOVA','PIZZA PIZZA',
    'PIZZAVILLE','LITTLE CAESAR','PAPA JOHN','DOMINO','PIZZA HUT','241 PIZZA',
    'MUCHO BURRITO','BAR BURRITO','BURRITO BOYZ',
    'THAI EXPRESS','EDO JAPAN','BENTO BENTO','FRESHII','BOOSTER JUICE',
    'SECOND CUP','SMOKE\'S POUTINERIE','SMOKES POUTINERIE',
    'HERO BURGER','HERO CERTIFIED','FIVE GUYS','NEW YORK FRIES',
    'CHIPOTLE','TACO BELL','TACO TIME',
    'DAIRY QUEEN','BASKIN-ROBBIN','BASKIN ROBBIN',
    'SWISS CHALET','ST-HUBERT','WHITE SPOT',
    'DOLLARAMA','SHOPPERS DRUG MART','7-ELEVEN','CIRCLE K','COUCHE-TARD',
    'FRESHCO','METRO','SOBEYS','LOBLAWS','NO FRILLS','COSTCO','WALMART',
    'BENTO SUSHI', 'FAT BASTARD BURRITO',
)
import re as _re_chain
def _is_chain(name_upper):
    """Start-of-name match only — avoids false positives like 'X & THAI EXPRESS' being
    matched as the 'THAI EXPRESS' chain."""
    n = (name_upper or '').strip()
    for c in _CHAIN_DENYLIST:
        if _re_chain.match(r'^' + _re_chain.escape(c) + r'(\b|$|[/#@,])', n):
            return True
    return False

def _resolve_cuisine(name_upper, name_raw, address):
    """Returns (cuisine_key, source) or (None, None) to drop the entry.
    Priority: chain denylist > web_verify (search-informed) > LLM name-only > keyword.
    """
    if _is_chain(name_upper):
        return None, None
    key = f"{name_upper}||{address.upper()}"
    w = _web_verify_cache.get(key)
    if w and w.get('status') == 'ok' and w.get('cuisine'):
        c = w['cuisine']
        if c == 'unknown': return None, None
        if c in _VALID_LLM_CUISINES: return c, 'web_search'
    llm = _llm_cache.get(key)
    if llm and llm.get('status') == 'ok':
        c = llm.get('cuisine')
        if c == 'unknown': return None, None
        if c in _VALID_LLM_CUISINES: return c, 'llm'
    for cuisine_k, keys in CUISINE_PATTERNS.items():
        for kw in keys:
            if kw in name_upper: return cuisine_k, 'keyword'
    return None, None

# Load Places cache up front so we can gate entries to verified-OPERATIONAL only
_PLACES_PATH = ROOT / 'tools' / 'cache' / 'places_cache.json'
_places = json.loads(_PLACES_PATH.read_text()) if _PLACES_PATH.exists() else {}
def _places_lookup(name, address):
    k = f"{(name or '').strip().upper()}||{(address or '').strip().upper()}"
    p = _places.get(k)
    if not p or p.get('status') != 'ok': return None
    return p

opens_by_cuisine = {}
n_food_active = 0; n_food_active_365 = 0; n_tagged_365 = 0; n_tagged_30 = 0
n_drop_unverified = 0; n_drop_closed = 0
_seen_for_dedup = {}   # (name_upper, addr_upper) → entry with earliest issuedDate
with open('/tmp/business_licences_alt.csv', encoding='utf-8', errors='replace') as f:
    rdr = csv.DictReader(f)
    for row in rdr:
        cat = (row.get('Category') or '').strip()
        if cat not in FOOD_CATS_FOR_CUISINE: continue
        if (row.get('Cancel Date') or '').strip(): continue
        n_food_active += 1
        iss = parse_d(row.get('Issued'))
        if not iss or iss < WINDOW_365: continue
        n_food_active_365 += 1
        op_raw = (row.get('Operating Name') or '').strip()
        op = op_raw.upper()
        if not op: continue
        addr1 = (row.get('Licence Address Line 1') or '').strip()
        addr3 = (row.get('Licence Address Line 3') or '').strip()
        addr_full = (addr1 + ' ' + addr3).strip()
        matched, source = _resolve_cuisine(op, op_raw, addr_full)
        if not matched: continue
        # Verification gate: only include Places-OPERATIONAL entries
        pdata = _places_lookup(op_raw, addr_full)
        if pdata is None:
            n_drop_unverified += 1; continue
        if pdata.get('businessStatus') != 'OPERATIONAL':
            n_drop_closed += 1; continue
        entry = {
            'operatingName': op_raw,
            'cuisine': matched,
            'cuisineSource': source,
            'issuedDate': iss.strftime('%Y-%m-%d'),
            'daysOpen': max(0, (TODAY - iss).days),
            'address': addr1,  # street only; postal kept in addr_full for cache lookups
            'businessStatus': pdata.get('businessStatus'),
        }
        for k in ('website', 'mapsUrl', 'rating', 'reviewCount', 'matchedName', 'lat', 'lng'):
            if pdata.get(k) is not None: entry[k] = pdata[k]
        # Dedupe by (name, address) — keep earliest issuedDate (true opening, not renewal)
        dkey = (op_raw.upper(), addr1.upper())
        existing = _seen_for_dedup.get(dkey)
        if existing is None or entry['issuedDate'] < existing['issuedDate']:
            _seen_for_dedup[dkey] = entry

# Build by-cuisine + counts from deduped set
for entry in _seen_for_dedup.values():
    n_tagged_365 += 1
    if entry['daysOpen'] <= 30: n_tagged_30 += 1
    opens_by_cuisine.setdefault(entry['cuisine'], []).append(entry)
for c in opens_by_cuisine:
    opens_by_cuisine[c].sort(key=lambda r: r['issuedDate'], reverse=True)
no_cuisines = []
for c, entries in opens_by_cuisine.items():
    no_cuisines.append({
        'key': c, 'label': CUISINE_LABEL.get(c, c),
        'count365d': len(entries),
        'count30d': sum(1 for e in entries if e['daysOpen'] <= 30),
        'newest': entries[0],
        'recent5': entries[:10],
    })
no_cuisines.sort(key=lambda r: -r['count365d'])
no_recent = []
for c, entries in opens_by_cuisine.items(): no_recent.extend(entries)
no_recent.sort(key=lambda r: r['issuedDate'], reverse=True)
out['newOpenings'] = {
    'asOf': TODAY.strftime('%Y-%m-%d'),
    'windowDays': 365,
    'totalActiveScanned': n_food_active,
    'totalNewActive365d': n_food_active_365,
    'totalTagged365d': n_tagged_365,
    'totalTagged30d': n_tagged_30,
    'tagRate365d': round(n_tagged_365 / max(n_food_active_365, 1) * 100, 1),
    'cuisines': no_cuisines,
    'recent': no_recent[:300],
}
log(f"  new openings: {n_tagged_365} tagged & verified in 12mo, {n_tagged_30} in 30d, across {len(no_cuisines)} cuisines (dropped {n_drop_unverified} unverified + {n_drop_closed} closed)")

# ---------------------------------------------------------------
# RECENT DEV APPLICATIONS FEED — last 90 days, in any corridor
# ---------------------------------------------------------------
log("Building dev apps feed…")
dev_feed = []
WINDOW_90D_DEV = TODAY - timedelta(days=180)  # widen to 6 months for narrative thickness
with open('/tmp/dev_apps.csv', encoding='utf-8', errors='replace') as f:
    rdr = csv.DictReader(f)
    for row in rdr:
        d = parse_iso(row.get('DATE_SUBMITTED'))
        if not d or d < WINDOW_90D_DEV: continue
        try: x = float(row['X']); y = float(row['Y'])
        except: continue
        lng, lat = UTM_TO_WGS84.transform(x, y)
        p = Point(lng, lat)
        matched_slug = None
        for c in CORRIDORS:
            if c['_shape'] is None: continue
            bx0, by0, bx1, by1 = c['_bbox']
            if not (bx0 <= lng <= bx1 and by0 <= lat <= by1): continue
            if c['_shape'].contains(p): matched_slug = c['slug']; break
        if not matched_slug: continue
        addr_parts = [row.get('STREET_NUM') or '', row.get('STREET_NAME') or '', row.get('STREET_TYPE') or '']
        addr = ' '.join(p for p in addr_parts if p).strip()
        dev_feed.append({
            'address': addr,
            'corridor': matched_slug,
            'applicationType': (row.get('APPLICATION_TYPE') or '').strip(),
            'status': (row.get('STATUS') or '').strip(),
            'date': d.strftime('%Y-%m-%d'),
            'description': ((row.get('DESCRIPTION') or '').strip())[:240],
            'ward': (row.get('WARD_NAME') or '').strip(),
            'contact': (row.get('CONTACT_NAME') or '').strip(),
        })
dev_feed.sort(key=lambda r: r['date'], reverse=True)
out['recentDevApps'] = dev_feed[:30]
log(f"  {len(dev_feed)} dev apps in 12 corridors over last 6 months; keeping latest 30")

# ---------------------------------------------------------------
# HEADLINE FINDINGS — computed insights for the homepage
# (placed AFTER risk-index + ranks attachment so all per-corridor fields exist)
# ---------------------------------------------------------------
# Make sure activeDevApps12m, etc. are attached to out['corridors'] before reading.
for i, c in enumerate(out['corridors']):
    if 'activeDevApps12m' not in c:
        c['activeDevApps12m'] = CORRIDORS[i]['_dev_apps']
    if 'majorCrime12m' not in c:
        c['majorCrime12m'] = CORRIDORS[i]['_mci']
    if 'storefrontCancellations90d' not in c:
        c['storefrontCancellations90d'] = CORRIDORS[i]['_cancellations_90d']
    if 'personsInCrisis12m' not in c:
        c['personsInCrisis12m'] = CORRIDORS[i]['_pic']

by_risk = sorted([c for c in out['corridors'] if c['parcels'] >= 5], key=lambda c: -(c.get('riskIndex') or 0))
by_upside = sorted([c for c in out['corridors'] if c['parcels'] >= 5], key=lambda c: -(c.get('developerUpsideCAD') or 0))
by_heritage_low = sorted([c for c in out['corridors'] if c['parcels'] >= 100], key=lambda c: (c.get('heritageDesignated') or 0))
by_heritage_high = sorted([c for c in out['corridors'] if c['parcels'] >= 100 and (c.get('heritageDesignated') or 0) > 0], key=lambda c: -(c.get('heritageDesignated') or 0))
by_cancel = sorted([c for c in out['corridors'] if c['parcels'] >= 5], key=lambda c: -(c.get('storefrontCancellations90d') or 0))
by_devapps = sorted([c for c in out['corridors'] if c['parcels'] >= 5], key=lambda c: -(c.get('activeDevApps12m') or 0))

total_lost_90d = sum((c.get('storefrontCancellations90d') or 0) for c in out['corridors'])

# Spread findings across DIFFERENT corridors — avoid having one corridor dominate the page.
USED_CORRIDORS = set()
def first_unused(sorted_list, exclude=None):
    """Return the first corridor not yet used in a finding; falls back to first."""
    exclude = exclude or set()
    for c in sorted_list:
        if c['slug'] not in USED_CORRIDORS and c['slug'] not in exclude:
            return c
    return sorted_list[0] if sorted_list else None

# Finding 1: Heritage asymmetry — explicit ethnic-commercial-corridor contrast against well-protected one.
# Prefer a contrast between racialized commercial (low heritage) and protected (high heritage) — both ≥30 parcels.
hi_h = by_heritage_high[0]
# Pick lowest-heritage corridor that ALSO has a clearly-ethnic community (not Regent Park area)
ETHNIC_COMMERCIAL = {'little-jamaica','little-italy','little-portugal','little-india','corso-italia','greektown','koreatown','west-chinatown','east-chinatown','kensington-market','roncesvalles','parkdale'}
lo_h_candidates = [c for c in by_heritage_low if c['slug'] in ETHNIC_COMMERCIAL]
lo_h = lo_h_candidates[0] if lo_h_candidates else by_heritage_low[0]
USED_CORRIDORS.add(hi_h['slug']); USED_CORRIDORS.add(lo_h['slug'])

# Finding 2: Highest displacement pressure — use #1 risk, OR #2 if it's already in used
risk_choice = first_unused(by_risk) or by_risk[0]
USED_CORRIDORS.add(risk_choice['slug'])

# Finding 3: Most dev apps — pick a corridor not yet used
dev_choice = first_unused(by_devapps) or by_devapps[0]
USED_CORRIDORS.add(dev_choice['slug'])

# Finding 4: Most cancellations — pick a corridor not yet used
cancel_choice = first_unused(by_cancel) or by_cancel[0]

findings = [
    {
        'headline': "Toronto's heritage system isn't asymmetric. It's hierarchical.",
        'body': f"{hi_h['title']} has {hi_h['heritageDesignated']} City-designated heritage properties on its {hi_h['parcels']} parcels. {lo_h['title']}: {lo_h['heritageDesignated']} on {lo_h['parcels']}. Affluent residential gets conservation districts. Racialized commercial gets demolition permits.",
        'kind': "asymmetry",
    },
    {
        'headline': "The cultural corridor most exposed to development",
        'body': f"{risk_choice['title']} carries a Risk Index of {risk_choice['riskIndex']:.0f} — among the highest displacement pressures in the city. ${risk_choice['developerUpsideCAD']/1e6:.0f}M in unused buildable envelope, {risk_choice.get('storefrontCancellations90d',0)} storefronts closed in the last 90 days, {risk_choice.get('heritageDesignated',0)} heritage protections.",
        'kind': "leader",
    },
    {
        'headline': "Where developer money is filing in 2026",
        'body': f"{dev_choice['title']}: {dev_choice.get('activeDevApps12m',0)} new development applications filed in the last twelve months. The applicants are public record — names appear in the live ledger below.",
        'kind': "pressure",
    },
    {
        'headline': f"{total_lost_90d} storefronts closed in 90 days",
        'body': f"{cancel_choice['title']}: {cancel_choice.get('storefrontCancellations90d',0)} cancellations in the last quarter — among the fastest erosion in the city. Named businesses and ages-at-closure in the live ledger.",
        'kind': "loss",
    },
]
out['findings'] = findings
log(f"  {len(findings)} headline findings computed")

# ---------------------------------------------------------------
# PER-CORRIDOR LEGACY-BUSINESS NAMES (top 5 oldest still operating)
# ---------------------------------------------------------------
log("Building per-corridor legacy business lists…")
legacy_by_corridor = {slug: [] for slug in CORRIDOR_FSAS}
with open('/tmp/business_licences_alt.csv', encoding='utf-8', errors='replace') as f:
    rdr = csv.DictReader(f)
    for row in rdr:
        cat = (row.get('Category') or '').strip()
        if cat not in STOREFRONT_CATS: continue
        if (row.get('Cancel Date') or '').strip(): continue  # active only
        iss = parse_d(row.get('Issued'))
        if not iss: continue
        years = (TODAY - iss).days / 365.25
        if years < 15: continue  # 15+ years
        fsa = fsa_of(row.get('Licence Address Line 3'))
        if not fsa: continue
        for slug, fsas in CORRIDOR_FSAS.items():
            if fsa in fsas:
                legacy_by_corridor[slug].append({
                    'name': (row.get('Operating Name') or '').strip() or '(unnamed)',
                    'address': (row.get('Licence Address Line 1') or '').strip(),
                    'category': cat,
                    'issuedYear': iss.year,
                    'yearsOperating': round(years, 1),
                    'fsa': fsa,
                })
                break
# Sort by years operating descending, keep top 5
for slug in legacy_by_corridor:
    legacy_by_corridor[slug].sort(key=lambda r: -r['yearsOperating'])
    legacy_by_corridor[slug] = legacy_by_corridor[slug][:5]
# Attach to corridors
for c in out['corridors']:
    c['legacyBusinesses'] = legacy_by_corridor.get(c['slug'], [])
log(f"  legacy business lists built for {sum(1 for v in legacy_by_corridor.values() if v)} corridors")

# ---------------------------------------------------------------
# LEAD STORY — pick one dramatic specific business to anchor the homepage
# Criteria: oldest legacy business in the corridor with the LOWEST heritage protection
# (i.e., the corridor most likely to demolish a 50+-year institution)
# ---------------------------------------------------------------
log("Choosing lead-story anchor…")
candidates = []
for c in out['corridors']:
    if c['parcels'] < 30: continue
    if (c.get('heritagePer100Parcels') or 0) > 5: continue  # only weakly-protected corridors
    for b in c.get('legacyBusinesses', []):
        if b['yearsOperating'] < 40: continue
        candidates.append({
            'business': b,
            'corridor': c,
            'devUpsidePerParcelCAD': (c['developerUpsideCAD'] / max(c['parcels'], 1)) if c['parcels'] else 0,
        })
candidates.sort(key=lambda x: -x['business']['yearsOperating'])
if candidates:
    lead = candidates[0]
    out['leadStory'] = {
        'businessName': lead['business']['name'],
        'address': lead['business']['address'],
        'category': lead['business']['category'],
        'issuedYear': lead['business']['issuedYear'],
        'yearsOperating': lead['business']['yearsOperating'],
        'corridorTitle': lead['corridor']['title'],
        'corridorCommunity': lead['corridor'].get('community', ''),
        'corridorSlug': lead['corridor']['slug'],
        'corridorHeritagePer100Parcels': lead['corridor'].get('heritagePer100Parcels', 0),
        'corridorHeritageDesignated': lead['corridor'].get('heritageDesignated', 0),
        'corridorParcels': lead['corridor'].get('parcels', 0),
        'corridorDevUpsideCAD': lead['corridor'].get('developerUpsideCAD', 0),
        'corridorCancellations90d': lead['corridor'].get('storefrontCancellations90d', 0),
        'parcelDevMathCAD': round(lead['devUpsidePerParcelCAD']),
    }
    log(f"  Lead: {lead['business']['name']} ({lead['business']['yearsOperating']}y, {lead['corridor']['title']})")
else:
    out['leadStory'] = None
    log(f"  No suitable lead story candidate")

# Pull quotes from existing Toronto journalism / scholarship
out['pullQuotes'] = [
    {'text': "Critics slam 'ridiculous' decision to exclude Little Jamaica from mandatory affordable housing plan",
     'source': "CBC News Toronto",
     'url': "https://www.cbc.ca/news/canada/toronto/critics-slam-ridiculous-decision-to-exclude-little-jamaica-from-mandatory-affordable-housing-plan-1.6219721"},
    {'text': "The 'Likkle but Talawa' Community: Little Jamaica, Toronto's Black Cultural District, is on the Verge of Disappearing",
     'source': "Society for the Study of Architecture in Canada (2024)",
     'url': "https://canada-architecture.org/the-likkle-but-talawa-community-little-jamaica-torontos-black-cultural-district-is-on-the-verge-of-disappearing-2/"},
    {'text': "The construction of the LRT has exposed anti-Black racism and heightened local socio-economic vulnerabilities, revealing the gentrifying effects of the Eglinton Crosstown's approach to transportation planning and urban policy.",
     'source': "The Impact of Transit Development on Racialized Neighbourhoods (York University)",
     'url': "https://yorkspace.library.yorku.ca/items/841653b4-afec-4d94-b285-4fdda3091ade"},
    {'text': "How Little Jamaica is preserving local culture with Toronto's first Black-led land trust.",
     'source': "The Green Line",
     'url': "https://thegreenline.to/stories/black-led-land-trust-little-jamaica/"},
]

# Attach new fields to the output corridors
for i, c in enumerate(out['corridors']):
    src = CORRIDORS[i]  # same order
    c['activeDevApps12m'] = src['_dev_apps']
    c['majorCrime12m'] = src['_mci']
    c['personsInCrisis12m'] = src['_pic']
    c['storefrontCancellations90d'] = src['_cancellations_90d']
    # Per-100-parcels normalization (proxy density)
    p = max(c['parcels'], 1)
    c['mciPer100Parcels'] = round(src['_mci'] * 100 / p, 1)
    c['picPer100Parcels'] = round(src['_pic'] * 100 / p, 1)
    c['devAppsPer100Parcels'] = round(src['_dev_apps'] * 100 / p, 1)
    c['cancellationsPer100Parcels'] = round(src['_cancellations_90d'] * 100 / p, 1)
    # Use the proper-spatial-join heritage count (supersedes the cached parcels.geojson value)
    c['heritageDesignated'] = src['_heritage_proper']
    c['heritageAddresses'] = src['_heritage_addresses']
    c['heritagePer100Parcels'] = round(c['heritageDesignated'] * 100 / p, 1)
    # Copy the new data layers (Census, Shelter, Apt, RentSafeTO, Biz breakdown) that attached after out was built
    c['census'] = src.get('_census')
    c['shelters'] = src.get('_shelters')
    c['apt'] = src.get('_apt')
    c['rentsafe'] = src.get('_rentsafe')
    c['biz'] = src.get('_biz')
    c['cuisine'] = src.get('_cuisine')

OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
OUT_PATH.write_text(json.dumps(out, indent=2))
log(f"Wrote {OUT_PATH} ({OUT_PATH.stat().st_size//1000} KB)")

# Print risk-index leaderboard
print()
print("=" * 80)
print("  RISK INDEX LEADERBOARD (100 = highest displacement pressure)")
print("=" * 80)
sorted_corridors = sorted(out['corridors'], key=lambda c: -c['riskIndex'])
print(f"  {'#':>2}  {'CORRIDOR':<40} {'INDEX':>8} {'PER-PARCEL HERITAGE':>22}")
print("-" * 80)
for i, c in enumerate(sorted_corridors, 1):
    print(f"  {i:>2}  {c['title'][:39]:<40} {c['riskIndex']:>8.1f} {c['heritagePerParcel']:>22.3f}")

# Print headline
print()
print("=" * 80)
print(f"  TORONTO'S 12 CULTURAL CORRIDORS")
print(f"  Total parcels analysed: {out['totals']['parcels']:,}")
print(f"  Developer upside:       ${out['totals']['developerUpsideCAD']/1e9:.2f} BILLION")
print(f"  Embodied carbon at risk: {out['totals']['embodiedCarbonTonnes']/1000:.1f} kilotonnes CO2")
print(f"  Carbon social cost:     ${out['totals']['carbonCostCAD']/1e6:.1f}M")
print("=" * 80)
print()
print(f"  {'CORRIDOR':<40} {'PARCELS':>8} {'$ UPSIDE':>14} {'CARBON $':>12}")
print("-" * 80)
for c in out['corridors']:
    print(f"  {c['title'][:39]:<40} {c['parcels']:>8,} ${c['developerUpsideCAD']/1e6:>11,.0f}M ${c['carbonCostCAD']/1e6:>9,.1f}M")
