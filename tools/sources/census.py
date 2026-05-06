"""Compute `builtYear` (median dwelling vintage) and `existing` (occupied dwelling count)
per neighborhood from the Toronto Neighbourhood Profiles 2021 158-model XLSX.

See `tools/README.md` § Census Source for the locked schema (sheet name, row labels,
bracket-to-midpoint table, alias map).
"""

import logging
import statistics
import sys
import time
from pathlib import Path

import requests
from openpyxl import load_workbook

from .neighborhoods import Neighborhood

PACKAGE_ID = "neighbourhood-profiles"
RESOURCE_ID = "19d4a806-7385-4889-acf2-256f1e079060"
RESOURCE_URL = (
    "https://ckan0.cf.opendata.inter.prod-toronto.ca/dataset/"
    "6e19a90f-971c-46b3-852c-0c48c436d1fc/resource/"
    f"{RESOURCE_ID}/download/nbhd_2021_census_profile_full_158model.xlsx"
)
CACHE_FILENAME = "npp_2021.xlsx"
SHEET_NAME = "hd2021_census_profile"

EXISTING_LABEL = (
    "Total - Occupied private dwellings by structural type of dwelling - 25% sample data"
)
PERIOD_UNIVERSE_LABEL = (
    "Total - Occupied private dwellings by period of construction - 25% sample data"
)

# Bracket label → midpoint year. Order matters: oldest → newest, for the cumulative scan
# in `_resolve_built_year`. The 1955 midpoint for the open-ended "1960 or before" bracket
# is a pragmatic stand-in (see README); only used by neighborhoods whose median dwelling
# predates 1961.
BRACKET_MIDPOINTS: list[tuple[str, int]] = [
    ("1960 or before", 1955),
    ("1961 to 1980", 1970),
    ("1981 to 1990", 1985),
    ("1991 to 2000", 1995),
    ("2001 to 2005", 2003),
    ("2006 to 2010", 2008),
    ("2011 to 2015", 2013),
    ("2016 to 2021", 2018),
]

# Cosmetic punctuation/whitespace differences between XLSX header names and the canonical
# AREA_NAME values from the neighborhoods GeoJSON. Same neighborhoods — not fallbacks.
NAME_ALIASES: dict[str, str] = {
    "Cabbagetown-South St. James Town": "Cabbagetown-South St.James Town",
    "Danforth-East York": "Danforth East York",
    "East End Danforth": "East End-Danforth",
    "North St. James Town": "North St.James Town",
    "O`Connor Parkview": "O'Connor-Parkview",
    "Taylor Massey": "Taylor-Massey",
    "Yonge-St. Clair": "Yonge-St.Clair",
}

_log = logging.getLogger(__name__)


def _download_with_retries(url: str, dest: Path) -> None:
    backoffs = (0.5, 1.0, 2.0)
    for attempt in range(len(backoffs) + 1):
        try:
            with requests.get(url, stream=True, timeout=120) as r:
                r.raise_for_status()
                with dest.open("wb") as fp:
                    for chunk in r.iter_content(chunk_size=64 * 1024):
                        if chunk:
                            fp.write(chunk)
            return
        except (requests.RequestException, OSError) as e:
            if attempt == len(backoffs):
                raise
            wait = backoffs[attempt]
            _log.warning("download %s failed (attempt %d): %s — retrying in %ss",
                         url, attempt + 1, e, wait)
            time.sleep(wait)


def _ensure_cached(cache_dir: Path) -> Path:
    cache_dir.mkdir(parents=True, exist_ok=True)
    cached = cache_dir / CACHE_FILENAME
    if cached.exists() and cached.stat().st_size > 0:
        _log.info("using cached %s", cached)
        return cached
    _log.info("downloading census XLSX → %s", cached)
    _download_with_retries(RESOURCE_URL, cached)
    return cached


def _xlsx_header_to_canonical(xlsx_name: str, area_names: set[str]) -> str | None:
    if xlsx_name in area_names:
        return xlsx_name
    return NAME_ALIASES.get(xlsx_name)


def _resolve_built_year(universe: int | float | None,
                        bracket_counts: list[int | float | None]) -> int | None:
    """Return midpoint year of the bracket where cumulative share crosses 50%, or None.

    `universe` is the row 326 denominator (Total occupied private dwellings by period of
    construction); `bracket_counts` are the 8 per-bracket counts in BRACKET_MIDPOINTS
    order. Suppressed cells appear as 0, indistinguishable from a true zero per the
    README — None is treated identically to 0.
    """
    if not universe:
        return None
    cumulative = 0
    for count, (_label, midpoint) in zip(bracket_counts, BRACKET_MIDPOINTS):
        cumulative += count or 0
        if cumulative * 2 >= universe:
            return midpoint
    # Random-rounding (±20 per README risk #2) can leave cumulative just under universe.
    # Fall through to the newest bracket — the median is at or after this point.
    return BRACKET_MIDPOINTS[-1][1]


def compute_census(neighborhoods: list[Neighborhood], cache_dir: Path
                   ) -> tuple[dict[str, int], dict[str, int], list[str]]:
    """Returns `(built_year_by_name, existing_by_name, fallback_names)` keyed on
    AREA_NAME. Names whose data is missing or zero get a citywide-median `builtYear`
    and `existing = 0`, and are listed in `fallback_names`.
    """
    cached = _ensure_cached(Path(cache_dir))
    area_names = {n.name for n in neighborhoods}

    _log.info("opening %s with read_only=True", cached.name)
    wb = load_workbook(cached, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        raise RuntimeError(f"sheet {SHEET_NAME!r} not found in {cached.name}")
    ws = wb[SHEET_NAME]

    bracket_label_set = {b for b, _y in BRACKET_MIDPOINTS}
    needed_labels = bracket_label_set | {EXISTING_LABEL, PERIOD_UNIVERSE_LABEL}

    header_row: tuple | None = None
    rows_by_label: dict[str, tuple] = {}
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx == 1:
            header_row = row
            continue
        if row[0] is None:
            continue
        label = str(row[0]).strip()
        if label in needed_labels:
            rows_by_label[label] = row
            if len(rows_by_label) == len(needed_labels):
                break
    wb.close()

    if header_row is None:
        raise RuntimeError("XLSX has no header row")
    missing = needed_labels - rows_by_label.keys()
    if missing:
        raise RuntimeError(f"required row labels missing in census XLSX: {sorted(missing)}")

    col_to_canonical: dict[int, str] = {}
    unknown_xlsx_names: list[str] = []
    for col_idx, xlsx_name in enumerate(header_row):
        if col_idx == 0 or xlsx_name is None:
            continue
        canonical = _xlsx_header_to_canonical(str(xlsx_name), area_names)
        if canonical is None:
            unknown_xlsx_names.append(str(xlsx_name))
            continue
        col_to_canonical[col_idx] = canonical
    if unknown_xlsx_names:
        _log.warning("XLSX header names not in AREA_NAME and not aliased: %s",
                     unknown_xlsx_names)

    existing_row = rows_by_label[EXISTING_LABEL]
    period_universe_row = rows_by_label[PERIOD_UNIVERSE_LABEL]
    bracket_rows = [rows_by_label[label] for label, _y in BRACKET_MIDPOINTS]

    built_year_by_name: dict[str, int] = {}
    existing_by_name: dict[str, int] = {}
    needs_built_year_fallback: list[str] = []
    needs_existing_fallback: list[str] = []

    for col_idx, name in col_to_canonical.items():
        existing_val = existing_row[col_idx]
        existing_int = int(existing_val) if existing_val else 0
        if existing_int > 0:
            existing_by_name[name] = existing_int
        else:
            needs_existing_fallback.append(name)

        universe = period_universe_row[col_idx]
        bracket_counts = [r[col_idx] for r in bracket_rows]
        built_year = _resolve_built_year(universe, bracket_counts)
        if built_year is not None:
            built_year_by_name[name] = built_year
        else:
            needs_built_year_fallback.append(name)

    if built_year_by_name:
        median_year = int(statistics.median(built_year_by_name.values()))
    else:
        median_year = 1985
        _log.error("no built_year computed for any neighborhood — defaulting to %d",
                   median_year)
    for name in needs_built_year_fallback:
        built_year_by_name[name] = median_year
    for name in needs_existing_fallback:
        existing_by_name[name] = 0

    for nb in neighborhoods:
        if nb.name not in built_year_by_name:
            built_year_by_name[nb.name] = median_year
            needs_built_year_fallback.append(nb.name)
        if nb.name not in existing_by_name:
            existing_by_name[nb.name] = 0
            needs_existing_fallback.append(nb.name)

    fallback_names = sorted(set(needs_built_year_fallback) | set(needs_existing_fallback))
    _log.info("census: %d existing values, %d builtYear values, %d fallback names",
              sum(1 for v in existing_by_name.values() if v > 0),
              len(built_year_by_name),
              len(fallback_names))
    return built_year_by_name, existing_by_name, fallback_names


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, stream=sys.stderr,
                        format="%(asctime)s %(levelname)s %(message)s")
    from .neighborhoods import fetch_neighborhoods
    cache = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("tools/cache")
    nbs = fetch_neighborhoods(cache)
    by, ex, fb = compute_census(nbs, cache)
    print(f"built_year: {len(by)} entries; sample: {list(by.items())[:3]}")
    print(f"existing:   {len(ex)} entries; sample: {list(ex.items())[:3]}")
    print(f"fallbacks:  {len(fb)} {fb[:5]}{'...' if len(fb) > 5 else ''}")
